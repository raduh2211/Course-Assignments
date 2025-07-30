import pytest

from openpyxl import load_workbook


def get_test_data():
    test_data = []
    data_file = '.\TestData\LoginData.xlsx'

    testdata_workbook = load_workbook(data_file)
    ws = testdata_workbook.active

    # print(ws['A1'])
    # print(ws['A1'].value)
    # print(ws['B1'].value)

    for data in ws.iter_rows(values_only=True):
        print(data)

    return test_data

get_test_data()
