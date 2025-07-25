import time

import pytest
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook


def get_test_data():
    test_data = []
    data_file = '.\TestData\LoginData.xlsx'

    testdata_workbook = load_workbook(data_file)
    ws = testdata_workbook.active

    for data in ws.iter_rows(values_only=True):
        test_data.append(data)

    return test_data

@pytest.mark.parametrize('username,password',get_test_data())
def test_param(username,password):
    driver = webdriver.Chrome()

    time.sleep(5)

    driver.get('https://opensource-demo.orangehrmlive.com/web/index.php/auth/login')
    time.sleep(5)

    driver.find_element(By.NAME, 'username').send_keys(username)
    driver.find_element(By.NAME, 'password').send_keys(password)
    driver.find_element(By.XPATH,'//button').click()

    time.sleep(5)
    driver.close()



