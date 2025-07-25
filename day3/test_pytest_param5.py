import pytest

from openpyxl import load_workbook


def get_test_data():
    test_data = []
    data_file = '.\TestData\LoginData.xlsx'

    testdata_workbook = load_workbook(data_file)
    ws = testdata_workbook.active

    for data in ws.iter_rows(values_only=True):
        test_data.append(data)

    return test_data

@pytest.mark.parametrize('username,password',get_test_data())
def test_param(username,password):
    print(username)
    print(password)
