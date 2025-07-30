import openpyxl

import pytest


# fetching data from excel
def get_data():
    data=[]
    data_file = openpyxl.load_workbook("LoginData.xlsx")
    sheet = data_file.active

    for row in sheet.iter_rows(1,4,1,2,values_only=True):
        rowdata = tuple(row)
        data.append(rowdata)

    return data


@pytest.mark.parametrize("username,password",get_data())
def test_data_check(username,password):
    print(username+" "+password)
    data_file = openpyxl.load_workbook("LoginData.xlsx",read_only=False)


    sheet = data_file.active
    sheet['C1'] = 'passed'
    data_file.save("LoginData.xlsx")

