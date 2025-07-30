import openpyxl

import pytest

from openpyxl import load_workbook

def get_data():

   wb = load_workbook("testdata1.xlsx")
   ws = wb["Data"]  # or wb.active

   ws['C1'] = "Status"
   ws['C2'] = "Pass"
   ws['C3'] = "Fail"

   wb.save("testdata1.xlsx")

   # Get the first sheet (index 0)
   sheet = wb.worksheets[0]

   # Print sheet name
   print(sheet.title)

   # Get the first sheet (index 0)
   sheet = wb.worksheets[1]

   # Print sheet name
   print(sheet.title)


def test1():
   get_data()
   assert False


import logging

# log_obj = logging.getLogger()
# log_format = "%(levelname)s %(msg)s"
# logging.basicConfig(filename="logreports.txt", filemode="w",format=log_format, level="DEBUG")
# log_obj.debug('Log message - info')
# log_obj.info('Log message - info')
# log_obj.warning('Log message - info')
# log_obj.error('Log message - info')
# log_obj.critical('Log message - info')
